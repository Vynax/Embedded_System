import difflib
import tkinter as tk
import ntpath
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill
from tkinter import filedialog
from tkinter import simpledialog as sd
fileX = []
TextX = []


def path_leaf(path):
    head, tail = ntpath.split(path)
    return tail or ntpath.basename(head)


try:
    # 選取檔案
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilenames()

    # 準備excel檔
    wb = Workbook()
    # wb.create_sheet("result")
    #sheet = wb["result"]
    sheet = wb.active
    # 輸入檔名
    str_input = sd.askstring("彙整檔案名稱", "請輸入檔名")
    # 讀取檔案
    for index, filename in enumerate(file_path):
        fileX.append(open(filename, encoding="utf-8"))
        print(path_leaf(file_path[index]))
        # 寫入檔案名稱 openpyxl起始位置從1開始
        sheet.cell(row=index + 2, column=1).value = path_leaf(file_path[index])
        sheet.cell(row=1, column=index + 2).value = path_leaf(file_path[index])
        TextX.append(fileX[index].read())
        # 刪掉空白跟換行字元
        TextX[index] = TextX[index].replace(" ", "").replace("\n", "")

    # print(file_path)

    # 計算相似度
    for index, t in enumerate(TextX):
        for j in range(index + 1, len(TextX)):
            m = difflib.SequenceMatcher(None, t, TextX[j])
            r = round(m.ratio(), 3)
            print(str(r) + " ", end="")
            sheet.cell(row=index + 2, column=j + 2).value = r
            # 如果相似度大於80%就把該欄位塗成紅色
            if r >= 0.8:
                sheet.cell(row=index + 2, column=j + 2).fill = PatternFill(fill_type="solid",
                                                                           start_color='FFFF0000',
                                                                           end_color='FFFF0000')
        print()
    """for index, f in enumerate(file_path):
        with open(f, 'r') as foo:
            #print("hi")
            pass"""

finally:
    wb.save(str_input+".xlsx")
    for f in fileX:
        f.close()
